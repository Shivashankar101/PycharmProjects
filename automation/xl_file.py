import openpyxl
from openpyxl import Workbook

#workbook = Workbook()
path = "/Users/shiv/desktop/mercury.xlsx"
workbook = openpyxl.load_workbook(path)

sheet1 = workbook.active
#sheet2 = workbook.get_sheet_by_name("sheet2")

rows = sheet1.max_row
cols = sheet1.max_column

print(rows)
print(cols)

for r in range(5, 8):
    for c in range(1, cols+1):
        sheet1.cell(row=r, column=c).value == "its okay"

workbook.save(path)


# for r in range(1, rows+1):
#     for c in range(1, cols+1):
#         print(sheet1.cell(row=r, column=c).value, end="     ")
#
#     print()

